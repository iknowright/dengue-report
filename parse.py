import json
import logging
import re
import sys
from datetime import date, datetime, timedelta
from io import BytesIO
from pprint import pprint

import boto3
import twd97
from openpyxl import load_workbook


def generate_week_str(now_week_start, week_ago_num):
    week_ago_start = now_week_start - timedelta(days=7 * week_ago_num)
    week_ago_end = week_ago_start + timedelta(days=6)
    return '%s~%s' % (
        week_ago_start.strftime("%Y-%m-%d"), week_ago_end.strftime("%Y-%m-%d"))


log = logging.getLogger('')
log.setLevel(logging.DEBUG)
format = logging.Formatter(
    "%(asctime)s - %(name)s - %(levelname)s - %(message)s")

ch = logging.StreamHandler(sys.stdout)
ch.setFormatter(format)
log.addHandler(ch)

parselogger = logging.getLogger('parse')
sheetlogger = logging.getLogger('sheet')
progresslogger = logging.getLogger('progress')

s3 = boto3.resource('s3')
s3_client = boto3.client('s3')
file_list = list()

weeknum = int(datetime.now().strftime('%U'))
this_year_str = datetime.now().strftime('%Y')
accept_weeks = []
this_week_start = date.today() - timedelta(days=date.today().weekday() + 1)
this_week_str = generate_week_str(this_week_start, 0)
last_week_str = generate_week_str(this_week_start, 1)
update_weeks =  [this_week_str, last_week_str]

# Updating 2 weeks data, which avg_egg_count need 5 week to calc, so accept total 6 weeks data
accept_weeks.append(str(weeknum))
accept_weeks.append(str(weeknum - 1))
accept_weeks.append(str(weeknum - 2))
accept_weeks.append(str(weeknum - 3))
accept_weeks.append(str(weeknum - 4))
accept_weeks.append(str(weeknum - 5))

parselogger.info('update_weeks: ' + str(update_weeks))
parselogger.info('accept_week: ' + str(accept_weeks))

for key in s3.Bucket('dengue-report-source').objects.all():
    if key.key.endswith(".xlsx"):
        if len(key.key.split("/")) < 2:
            continue
        city = key.key.split("/")[0]
        file_name = key.key.split("/")[1]
        file_list.append({
            "city": city,
            "file_name": file_name,
            "file_key": key.key
        })

bucket_dict = dict()
survey_dict = dict()
for file_dict in file_list:
    # s3_client.download_file(
        # 'dengue-report-source',
        # file_dict['file_key'],
        # file_dict['file_name']
    # )

    # wb = load_workbook(file_dict['file_name'], read_only=True)
    s3_obj = s3.Object('dengue-report-source', file_dict['file_key'])
    wb = load_workbook(filename=BytesIO(s3_obj.get()['Body'].read()), read_only=True)
    parselogger.info(file_dict['file_name'])
    city = file_dict['city']
    for sheet_name in wb.get_sheet_names():
        sheet_name_match = re.search(r'(\d+)(年)?第(\d+)(週|周)', sheet_name)
        if sheet_name == '誘卵桶資訊':
            ws = wb['誘卵桶資訊']
            sheetlogger.info(sheet_name)
            progresslogger.info('Bucket handle begin')
            for row in range(3, ws.max_row+1):
                bucket_id = ws['A' + str(row)].value
                if bucket_id is None:
                    continue

                bucket_x = ws['B' + str(row)].value
                bucket_y = ws['C' + str(row)].value
                if bucket_x is None or bucket_y is None:
                    continue
                try:
                    bucket_x = float(bucket_x)
                    bucket_y = float(bucket_y)
                except:
                    continue

                bucket_address = ws['D' + str(row)].value
                bucket_note = ws['E' + str(row)].value
                bucket_inside_or_outside = ws['F' + str(row)].value

                if bucket_inside_or_outside is None or bucket_inside_or_outside == '外':
                    bucket_inside_or_outside = '外'

                bucket_lat, bucket_lng = twd97.towgs84(bucket_x, bucket_y)
                bucket_dict[bucket_id] = {
                    'bucket_lat': bucket_lat,
                    'bucket_lng': bucket_lng,
                    'bucket_inside_or_outside': bucket_inside_or_outside,
                    # 'bucket_address': bucket_address,
                    # 'bucket_note': bucket_note
                }
            progresslogger.info('finish ' + str(ws.max_row + 1) + ' buckets')
        elif sheet_name_match and sheet_name_match.group(3) in accept_weeks and sheet_name_match.group(1) == this_year_str:
            ws = wb[sheet_name]
            sheetlogger.info(sheet_name)
            progresslogger.info('record handle begin')
            for row in range(3, ws.max_row+1):
                survey_date = ws['A' + str(row)].value
                bucket_id = ws['B'+ str(row)].value
                if (isinstance(survey_date, datetime) is False and
                    bucket_dict.get(bucket_id) is None):
                    break
                elif (isinstance(survey_date, datetime) is False or
                      bucket_dict.get(bucket_id) is None):
                    continue

                area = ws['C' + str(row)].value
                village = ws['D' + str(row)].value
                if '里' not in village:
                    village = village + '里'

                egg_num = ws['E' + str(row)].value if ws['E' + str(row)].value != None else '暫無資料'
                if egg_num == 0:
                    egypt_egg_num = 0
                    white_egg_num = 0
                else:
                    egypt_egg_num = ws['F' + str(row)].value if ws['F' + str(row)].value is not None else '暫無資料'
                    white_egg_num = ws['G'+ str(row)].value if ws['G' + str(row)].value is not None else '暫無資料'

                larvae_num = ws['H' + str(row)].value if ws['H' + str(row)].value is not None else '暫無資料'
                if larvae_num == 0:
                    egypt_larvae_num = 0
                    white_larvae_num = 0
                else:
                    egypt_larvae_num = ws['I' + str(row)].value if ws['I' + str(row)].value != None else '暫無資料'
                    white_larvae_num = ws['J' + str(row)].value if ws['J' + str(row)].value != None else '暫無資料'
                survey_note = ws['K' + str(row)].value if ws['K' + str(row)].value != None else '無'

                week_start = survey_date - timedelta(days=survey_date.weekday()+1)
                week_end = week_start + timedelta(days=6)
                week_range_str = '%s~%s' % (
                    week_start.strftime("%Y-%m-%d"), week_end.strftime("%Y-%m-%d"))

                if survey_dict.get(week_range_str) is None:
                    survey_dict[week_range_str] = dict()
                    survey_dict[week_range_str][city] = dict()
                    survey_dict[week_range_str][city][area] = dict()
                    survey_dict[week_range_str][city][area][village] = dict()
                elif survey_dict[week_range_str].get(city) is None:
                    survey_dict[week_range_str][city] = dict()
                    survey_dict[week_range_str][city][area] = dict()
                    survey_dict[week_range_str][city][area][village] = dict()
                elif survey_dict[week_range_str][city].get(area) is None:
                    survey_dict[week_range_str][city][area] = dict()
                    survey_dict[week_range_str][city][area][village] = dict()
                elif survey_dict[week_range_str][city][area].get(village) is None:
                    survey_dict[week_range_str][city][area][village] = dict()

                survey_dict[week_range_str][city][area][village][bucket_id] = {
                    "bucket_id": bucket_id,
                    "survey_date": survey_date.strftime("%Y-%m-%d"),
                    "egg_num": egg_num,
                    "egypt_egg_num": egypt_egg_num,
                    "white_egg_num": white_egg_num,
                    "larvae_num": larvae_num,
                    "egypt_larvae_num": egypt_larvae_num,
                    "white_larvae_num": white_larvae_num,
                    "survey_note": survey_note,
                }
            progresslogger.info('finish ' + str(ws.max_row + 1) + ' records')
    wb.close()

parselogger.info('avg_egg_num counting begin')
parselogger.info(str(survey_dict))
for week_range_str in update_weeks:
    week_start = datetime.strptime(week_range_str.split("~")[0], "%Y-%m-%d")
    week_str_list = list()
    week_str_list.append(generate_week_str(week_start, 1))
    week_str_list.append(generate_week_str(week_start, 2))
    week_str_list.append(generate_week_str(week_start, 3))
    week_str_list.append(generate_week_str(week_start, 4))

    for city in survey_dict[week_range_str].keys():
        for area in survey_dict[week_range_str][city].keys():
            for village in survey_dict[week_range_str][city][area].keys():
                for bucket_id in survey_dict[week_range_str][city][area][village].keys():
                    total_egg_num = 0
                    for week_str in week_str_list:
                        try:
                            total_egg_num = total_egg_num + \
                                survey_dict[week_str][city][area][village][bucket_id]
                        except:
                            continue
                    survey_dict[week_range_str][city][area][village][bucket_id]['avg_egg_num'] = int(total_egg_num / len(week_str_list))
parselogger.info('avg_egg_num counting end')

s3.Object("dengue-report-dest", "bucket-list.json").put(
    ACL='public-read',
    Body=json.dumps(bucket_dict, indent=4, ensure_ascii=False)
)
parselogger.info('upload s3 begin')
for week_range_str in update_weeks:
    s3.Object("dengue-report-dest", "week/%s.json" % (week_range_str)).put(
        ACL='public-read',
        Body=json.dumps(survey_dict[week_range_str], ensure_ascii=False)
    )
    # DEBUGGING
    # with open('%s.json' % (week_range_str), 'w') as myfile:
    #     json.dump(
    #         survey_dict[week_range_str],
    #         myfile,
    #         indent=4,
    #         ensure_ascii=False
    #     )
parselogger.info('upload s3 end')
